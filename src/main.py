import asyncio
import os
import random
import sqlite3
from datetime import datetime

import openpyxl
from telebot import types
from telebot.async_telebot import AsyncTeleBot

TOKEN = os.environ.get("TOKEN")

if not TOKEN:
    raise ValueError("Токен не найден")

bot = AsyncTeleBot(TOKEN)

congratulations = [
    "Поздравляем с днем рождения! Желаем здоровья, удачи, любви, везения, мира, добра, улыбок, благополучия. Пусть все мечты исполняются. Пусть жизнь будет долгой и гладкой, полной ярких и запоминающихся событий!",
    "Хочу пожелать тебе не просто благополучия и хорошей жизни, а еще и удачного стечения обстоятельств. Чтобы достижениями своими ты мог гордиться. Живи ярко и улыбайся каждый день!",
    "С днем рождения! Желаю добра, света, мира, улыбок, отличного настроения. Пусть всё плохое обходит стороной, жизненные невзгоды преодолеваются с легкостью, а каждый день будет наполнен радостью и счастьем. И конечно, светлой веры, огромной надежды, бесконечной любви.",
    "В день рождения желаю легкости чувств, душевной бодрости, ясности мыслей и ярких позитивных ощущений. Пусть жизненная энергия поможет достичь самых невероятных замыслов.",
    "Поздравляю с днем рождения! Пусть счастье сопровождает вас на каждом пути, здоровье оберегает при любых обстоятельствах. Удача и везение станут вашим надежным спутником по жизни. Желаю вдохновения во всем, позитивного настроения и исполнения всех планов."
    "Желаю здоровья и хорошего настроения, всех благ и удовольствий жизни, благополучия и домашнего уюта, любви и человеческого счастья!"
    "Пусть в твой день рождения все будет необыкновенным и замечательным, словно в волшебных сказках, случаются чудеса, а счастье будет прекрасном, как радуга!"
    "Желаю самые приятные минуты жизни разделить с любимым человеком, радоваться мгновениям волшебного счастья, быть желанным и нужным в любой компании!"
    "Пусть свет счастливой звезды ведет тебя навстречу радостным и интересным событиям, радужным надеждам и мечтам, к успеху и процветанию!"
    "Желаем уютной атмосферы в доме, любви и теплоты в отношениях, уважения и доверия в коллективе, счастливых и радостных лет жизни!"
    "Пусть добрый художник раскрашивает твою жизнь лишь светлыми красками, а дни приносят впечатления, которые хочется вспоминать. С днем рождения!",
]


@bot.message_handler(commands=["start"])
async def handle_start_command(message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
    keyboard.add(
        types.KeyboardButton("Запрос"),
        types.KeyboardButton("Получить идею для поздравления"),
        types.KeyboardButton("Создать базу данных"),
        types.KeyboardButton(
            "Отправить файл Excel"
        ),  # Добавлена кнопка "Отправить файл Excel"
    )
    await bot.send_message(
        message.chat.id, "Привет! Выберите действие:", reply_markup=keyboard
    )


@bot.message_handler(content_types=["text"])
async def handle_text(message):
    text = message.text.strip()

    if text == "Запрос":
        await handle_birthday_request(message)

    elif text == "Получить идею для поздравления":
        congratulation = random.choice(congratulations)
        await bot.send_message(message.chat.id, congratulation)

    elif text == "Создать базу данных":
        await create_database_table()
        await clear_database(message)
        await bot.send_message(message.chat.id, "База данных успешно создана!")

    elif text == "Отправить файл Excel":
        await bot.send_message(message.chat.id, "Пожалуйста, отправьте файл Excel.")


@bot.message_handler(content_types=["document"])
async def handle_document(message):
    if (
        message.document.mime_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        await export_to_sqlite(message)
        await bot.send_message(
            message.chat.id,
            "Файл Excel успешно обработан и данные добавлены в базу данных.",
        )
    else:
        await bot.send_message(
            message.chat.id, "Пожалуйста, отправьте файл в формате Excel."
        )


async def handle_birthday_request(message):
    current_date = datetime.now().strftime("%d.%m")
    database_path = os.path.join(os.getcwd(), "People.sqlite3")
    connection = sqlite3.connect(database_path)
    cursor = connection.cursor()
    select_query = "SELECT * FROM People"

    try:
        cursor.execute(select_query)
        records = cursor.fetchall()
    except sqlite3.OperationalError:
        await bot.send_message(message.chat.id, "База данных пуста")
        return

    birthday_names = [record[0] for record in records if record[1][:5] == current_date]

    if birthday_names:
        birthday_message = "\n".join(
            f"{index + 1}. {name}" for index, name in enumerate(birthday_names)
        )
        await bot.send_message(
            message.chat.id,
            f"Сегодня дней рождений {len(birthday_names)}:\n{birthday_message}",
        )
    else:
        await bot.send_message(message.chat.id, "Сегодня дней рождений нет :(")


async def create_database_table():
    database_path = os.path.join(os.getcwd(), "People.sqlite3")
    connection = sqlite3.connect(database_path)
    cursor = connection.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS People (name text, date text)")
    connection.commit()
    cursor.close()
    connection.close()


async def clear_database(message: types.Message):
    database_path = os.path.join(os.getcwd(), "People.sqlite3")
    connection = sqlite3.connect(database_path)
    cursor = connection.cursor()
    cursor.execute("DELETE FROM People")
    connection.commit()
    cursor.close()
    connection.close()


async def export_to_sqlite(message: types.Message):
    database_path = os.path.join(os.getcwd(), "People.sqlite3")
    connection = sqlite3.connect(database_path)
    cursor = connection.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS People (name text, date text)")

    file_info = await bot.get_file(message.document.file_id)
    file_path = file_info.file_path
    downloaded_file = await bot.download_file(file_path)

    with open("temp.xlsx", "wb") as new_file:
        new_file.write(downloaded_file)

    workbook = openpyxl.load_workbook("temp.xlsx", data_only=True)
    sheet = workbook.active
    for row in range(1, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        date = sheet.cell(row, 2).value
        cursor.execute("INSERT INTO People VALUES (?, ?);", (name, date))

    connection.commit()
    cursor.close()
    connection.close()
    os.remove("temp.xlsx")


asyncio.run(bot.polling(non_stop=True))
